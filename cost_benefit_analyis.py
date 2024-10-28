import tkinter as tk
from tkinter import ttk
from tkinter.scrolledtext import ScrolledText
from tkinter import messagebox, filedialog
import pandas as pd
import re
import tkinter.font as tkFont
import os

# Import matplotlib modules for charting
import matplotlib
matplotlib.use("TkAgg")  # Use TkAgg backend for Tkinter
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

# ----------------------------------------
# Scrollable Frame Class
# ----------------------------------------

class ScrollableFrame(ttk.Frame):
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)
        
        canvas = tk.Canvas(self, borderwidth=0, background="#f0f0f0")
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        # Use tk.Frame instead of ttk.Frame to allow setting background
        self.scrollable_frame = tk.Frame(canvas, background="#f0f0f0")
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )
        
        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

# ----------------------------------------
# Main Application Window
# ----------------------------------------

# Create the main application window
root = tk.Tk()
root.title('Cost-Benefit Analysis Toolkit')
root.geometry('900x700')  # Increased window size to accommodate charts

# Initialize ttk.Style and set the 'clam' theme
style = ttk.Style()
style.theme_use('clam')

# Configure the 'TButton' style
style.configure('TButton',
                background='#d0e8f1',
                foreground='black',
                padding=6)
style.map('TButton',
          background=[('active', '#87CEFA')],
          foreground=[('active', 'black')])

# Create a Scrollable Frame to hold the Notebook
scrollable_container = ScrollableFrame(root)
scrollable_container.pack(fill='both', expand=True)

# Create a Notebook widget to hold tabs within the scrollable frame
notebook = ttk.Notebook(scrollable_container.scrollable_frame)
notebook.pack(expand=1, fill='both')

# ----------------------------------------
# First Tab: About CBA
# ----------------------------------------

# Create a frame for the 'About CBA' tab
about_cba_frame = ttk.Frame(notebook)
notebook.add(about_cba_frame, text='About CBA')

# Create a ScrolledText widget for the content
scrolled_text = ScrolledText(about_cba_frame, wrap='word', font=('Arial', 12))
scrolled_text.pack(expand=1, fill='both')

# Insert the formatted content
content = '''
1. **Define Scope:** Clarify procurement project goals, including evaluated products & desired outcomes.

2. **Identify Costs:**
   - **Direct Costs:** Purchase price, shipping, clearance.
   - **Indirect Costs:** Maintenance, training, potential downtime.
   - **Opportunity Costs:** Evaluate potential revenue loss from procurement choices.

3. **Identify Benefits:**
   - **Tangible Benefits:** Measure increased revenue, cost savings, efficiency, quality improvements.
   - **Intangible Benefits:** Consider non-quantifiable gains like brand reputation & customer satisfaction.

4. **Data Collection:** Collect data from past purchases, supplier quotes, market research.

5. **Quantify Costs & Benefits:**
   - Convert costs & benefits into a common monetary unit for comparison.
   - Estimate future cash flows & discount them to present value.

6. **Analyze Data:**
   - Calculate total costs & benefits.
   - Determine net present value (NPV) by subtracting costs from benefits.
   - Evaluate additional metrics like benefit-cost ratio (BCR) & payback period.

7. **Sensitivity Analysis:**
   - Identify risks & uncertainties.
   - Assess the impact of changes in key assumptions on results.

8. **Document Findings:** Prepare a report detailing methods, assumptions, & results, supported by charts & graphs.

9. **Make Recommendations:** Decide on procurement, explore alternatives, & negotiate terms.

10. **Monitor Post-Decision:** Continuously track implementation & performance against anticipated benefits & costs to ensure alignment with goals.
'''

# Clear any existing content
scrolled_text.delete('1.0', tk.END)

# Insert the content
scrolled_text.insert('1.0', content)

# Apply formatting
# Define font styles
bold_font = ('Arial', 12, 'bold')
normal_font = ('Arial', 12)

# Tag configurations
scrolled_text.tag_configure('bold', font=bold_font, foreground='dark blue')
scrolled_text.tag_configure('normal', font=normal_font)
scrolled_text.tag_configure('indent', lmargin1=25, lmargin2=25)
scrolled_text.tag_configure('sub_indent', lmargin1=50, lmargin2=50)

# Apply tags

# Patterns for headings and subheadings
heading_pattern = r'^(\d+\.\s\*\*)(.+?)(\*\*)(.*)'
subheading_pattern = r'^(\s*-\s\*\*)(.+?)(\*\*)(.*)'

lines = content.split('\n')
current_index = 1.0

for line in lines:
    line = line.rstrip()
    if line.strip() == '':
        current_index += 1
        continue

    # Check for heading
    heading_match = re.match(heading_pattern, line)
    subheading_match = re.match(subheading_pattern, line)

    if heading_match:
        # Apply bold and color to heading
        start_idx = f"{current_index} + {len(heading_match.group(1))} chars"
        end_idx = f"{current_index} + {len(heading_match.group(1) + heading_match.group(2))} chars"
        scrolled_text.tag_add('bold', start_idx, end_idx)
        scrolled_text.tag_add('indent', f"{current_index}", f"{current_index} lineend")
    elif subheading_match:
        # Apply bold and color to subheading
        start_idx = f"{current_index} + {len(subheading_match.group(1))} chars"
        end_idx = f"{current_index} + {len(subheading_match.group(1) + subheading_match.group(2))} chars"
        scrolled_text.tag_add('bold', start_idx, end_idx)
        scrolled_text.tag_add('sub_indent', f"{current_index}", f"{current_index} lineend")
    else:
        # Normal text
        scrolled_text.tag_add('normal', f"{current_index}", f"{current_index} lineend")
        if line.startswith('   -'):
            scrolled_text.tag_add('sub_indent', f"{current_index}", f"{current_index} lineend")
        else:
            scrolled_text.tag_add('indent', f"{current_index}", f"{current_index} lineend")

    current_index += 1

# Disable editing
scrolled_text.config(state='disabled')

# Function to copy the content to clipboard
def copy_to_clipboard():
    try:
        scrolled_text.config(state='normal')
        content = scrolled_text.get('1.0', tk.END).strip()
        scrolled_text.config(state='disabled')
        root.clipboard_clear()
        root.clipboard_append(content)
        messagebox.showinfo("Copy Successful", "Content has been copied to the clipboard.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while copying to clipboard:\n{e}")

# Add the 'Copy' button below the scrolled text
copy_button_cba = ttk.Button(about_cba_frame, text="Copy", command=copy_to_clipboard)
copy_button_cba.pack(pady=10)

# ----------------------------------------
# Second Tab: BCR vs. Net Profit
# ----------------------------------------

# Add a new tab for 'BCR vs. Net Profit'
bcr_vs_profit_frame = ttk.Frame(notebook)
notebook.add(bcr_vs_profit_frame, text='BCR vs. Net Profit')

# Add a heading above the table
heading_label = ttk.Label(
    bcr_vs_profit_frame,
    text="Key Differences between BCR and Net Profit",
    font=('Arial', 14, 'bold')
)
heading_label.pack(pady=10)

# Create a frame for the scrollable table
table_container = ttk.Frame(bcr_vs_profit_frame)
table_container.pack(expand=1, fill='both', padx=10)

# Create a canvas inside the frame
table_canvas = tk.Canvas(table_container)
table_canvas.pack(side=tk.LEFT, fill='both', expand=1)

# Add scrollbars to the canvas
vsb = ttk.Scrollbar(table_container, orient="vertical", command=table_canvas.yview)
vsb.pack(side=tk.RIGHT, fill='y')

hsb = ttk.Scrollbar(bcr_vs_profit_frame, orient="horizontal", command=table_canvas.xview)
hsb.pack(side=tk.BOTTOM, fill='x')

table_canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

# Create a frame inside the canvas
table_frame = ttk.Frame(table_canvas)
table_canvas.create_window((0, 0), window=table_frame, anchor='nw')

# Define fonts
header_font = ('Arial', 10, 'bold')
cell_font = ('Arial', 10)
metric_color_font = ('Arial', 10)
metric_color = 'dark blue'

# Define columns and data for the table
columns = ("Metric", "BCR (Benefit-Cost Ratio)", "Net Profit")
data = [
    ("Definition:", "Ratio of PV of benefits to PV of costs", "Difference between total revenue and total costs"),
    ("Formula:", "PV of Benefits / PV of Costs", "Total Benefits - Total Costs"),
    ("Interpretation:", "Shows return per unit cost (BCR > 1 is viable)", "Shows absolute profit in monetary terms"),
    ("Focus:", "Relative comparison of benefits and costs", "Absolute profitability in currency"),
    ("Decision-Making:", "Useful for comparing projects", "Useful for assessing financial gain"),
    ("Time Value of Money:", "Uses discounted present values", "Does not require discounting")
]

# Calculate the maximum width for each column
def get_max_col_width(data_list, font):
    max_width = []
    for col in zip(*data_list):
        col_width = max([tkFont.Font(font=font).measure(str(item)) for item in col])
        max_width.append(col_width)
    return max_width

col_max_widths = get_max_col_width([columns] + data, cell_font)
total_width = sum(col_max_widths)

# Create the header labels
for col_index, col_name in enumerate(columns):
    header_label = ttk.Label(
        table_frame,
        text=col_name,
        font=header_font,
        anchor='center',
        borderwidth=1,
        relief='solid'
    )
    header_label.grid(row=0, column=col_index, sticky='nsew')
    # Set the column width
    table_frame.grid_columnconfigure(col_index, minsize=col_max_widths[col_index] + 20)

# List of Metrics to color
metrics_to_color = ["Definition", "Formula", "Interpretation", "Focus", "Decision-Making", "Time Value of Money"]

# Insert data into the table
for row_index, row_data in enumerate(data, start=1):
    for col_index, cell_value in enumerate(row_data):
        # Determine if this cell is in the Metric column and needs to be colored
        if col_index == 0 and cell_value.strip(':') in metrics_to_color:
            label = ttk.Label(
                table_frame,
                text=cell_value,
                font=metric_color_font,
                foreground=metric_color,
                anchor='w',
                borderwidth=1,
                relief='solid',
                wraplength=col_max_widths[col_index] + 20
            )
        else:
            label = ttk.Label(
                table_frame,
                text=cell_value,
                font=cell_font,
                anchor='w',
                borderwidth=1,
                relief='solid',
                wraplength=col_max_widths[col_index] + 20
            )
        label.grid(row=row_index, column=col_index, sticky='nsew')

    # Apply alternating row colors
    bg_color = "#f0f0f0" if row_index % 2 == 0 else "#ffffff"
    for col_index in range(len(columns)):
        label = table_frame.grid_slaves(row=row_index, column=col_index)[0]
        label.configure(background=bg_color)

# Adjust row weights
for row_index in range(len(data) + 1):
    table_frame.grid_rowconfigure(row_index, weight=1)

# Update the scroll region when the size of the frame changes
def on_frame_configure(event):
    table_canvas.configure(scrollregion=table_canvas.bbox("all"))

table_frame.bind("<Configure>", on_frame_configure)

# Function to download the table data to Excel
def download_to_excel_bcr():
    try:
        # Convert data to pandas DataFrame
        df = pd.DataFrame(data, columns=columns)
        # Save to Excel file at the specified path
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        file_path = os.path.join(desktop_path, 'BCR_vs_Net_Profit.xlsx')
        df.to_excel(file_path, index=False)
        messagebox.showinfo("Download Successful", f"Table data has been exported to 'BCR_vs_Net_Profit.xlsx' on your Desktop.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while exporting to Excel:\n{e}")

# Add the 'Download' button below the table
download_button_bcr = ttk.Button(bcr_vs_profit_frame, text="Download", command=download_to_excel_bcr)
download_button_bcr.pack(pady=10)

# ----------------------------------------
# Third Tab: Payback Period Calculator
# ----------------------------------------

# Add a new tab for 'Payback Period Calculator'
payback_period_frame = ttk.Frame(notebook)
notebook.add(payback_period_frame, text='Payback Period')

# Define the PaybackPeriodApp class adjusted for integration
class PaybackPeriodApp:
    def __init__(self, parent):
        self.parent = parent

        # Title Label
        ttk.Label(self.parent, text="Payback Period Calculator", font=("Helvetica", 16)).pack(pady=10)

        # Inputs for Initial Investment and Annual Net Benefits
        input_frame = ttk.Frame(self.parent)
        input_frame.pack(pady=10)

        ttk.Label(input_frame, text="Initial Investment (£): ").grid(row=0, column=0, padx=5, pady=5)
        self.initial_investment_entry = ttk.Entry(input_frame, width=15)
        self.initial_investment_entry.grid(row=0, column=1, padx=5, pady=5)
        self.initial_investment_entry.insert(0, "10000")  # Default Data

        # Instruction for Initial Investment
        ttk.Label(input_frame, text="(Enter as a positive number)").grid(row=0, column=2, padx=5, pady=5, sticky='w')

        ttk.Label(input_frame, text="Annual Net Benefits (£): ").grid(row=1, column=0, padx=5, pady=5)
        self.annual_cash_flow_entry = ttk.Entry(input_frame, width=15)
        self.annual_cash_flow_entry.grid(row=1, column=1, padx=5, pady=5)
        self.annual_cash_flow_entry.insert(0, "2500")  # Default Data

        # Update Button
        ttk.Button(self.parent, text="Update", command=self.update_table).pack(pady=10)

        # Table Frame
        self.table_frame = ttk.Frame(self.parent)
        self.table_frame.pack(pady=10, padx=20)

        # Set up Treeview (Table) with Scrollbar
        self.tree = ttk.Treeview(self.table_frame, columns=("Year", "Cash Flow", "Cumulative Cash Flow"), show="headings")
        self.tree.heading("Year", text="Year")
        self.tree.heading("Cash Flow", text="Cash Flow")
        self.tree.heading("Cumulative Cash Flow", text="Cumulative Cash Flow")
        self.tree.column("Year", anchor="center", width=50)
        self.tree.column("Cash Flow", anchor="center", width=100)
        self.tree.column("Cumulative Cash Flow", anchor="center", width=150)

        # Add vertical scrollbar
        scrollbar = ttk.Scrollbar(self.table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        self.tree.grid(row=0, column=0)
        scrollbar.grid(row=0, column=1, sticky="ns")

        # Calculate and Download Buttons
        button_frame = ttk.Frame(self.parent)
        button_frame.pack(pady=10)
        ttk.Button(button_frame, text="Calculate Payback Period", command=self.calculate_payback_period).grid(row=0, column=0, padx=10)
        ttk.Button(button_frame, text="Download to Excel", command=self.download_to_excel).grid(row=0, column=1, padx=10)

        # Payback Period Result Label
        self.result_label = ttk.Label(self.parent, text="", font=("Helvetica", 12))
        self.result_label.pack(pady=10)

        # Chart Frame
        self.chart_frame = ttk.Frame(self.parent)
        self.chart_frame.pack(pady=10, padx=20, fill='both', expand=True)

        # Initialize matplotlib Figure and Canvas
        self.figure = Figure(figsize=(6, 4), dpi=100)
        self.ax = self.figure.add_subplot(111)
        self.ax.set_title("Cumulative Cash Flow Over Time")
        self.ax.set_xlabel("Year")
        self.ax.set_ylabel("£")

        self.canvas = FigureCanvasTkAgg(self.figure, master=self.chart_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(fill='both', expand=True)

        # Download Buttons Frame for Chart
        download_buttons_frame = ttk.Frame(self.parent)
        download_buttons_frame.pack(pady=10)

        # Download Chart Button
        ttk.Button(download_buttons_frame, text="Download Chart", command=self.download_chart).grid(row=0, column=0, padx=10)

    def update_table(self):
        """Populate the table with cash flows and cumulative cash flows based on user input."""
        try:
            # Retrieve values from input boxes
            self.initial_investment = float(self.initial_investment_entry.get())
            self.annual_cash_flow = float(self.annual_cash_flow_entry.get())

            # Clear existing rows in the table
            self.tree.delete(*self.tree.get_children())

            # Populate table with new values, allowing for a longer payback period if needed
            cumulative_cash_flow = -self.initial_investment  # Start with negative initial investment
            year = 1
            max_years = 50  # Optional limit to prevent excessive calculation

            cash_flow_data = []

            while cumulative_cash_flow < 0 and year <= max_years:
                cumulative_cash_flow += self.annual_cash_flow
                cash_flow_display = f"£{self.annual_cash_flow:,.2f}"
                cumulative_display = f"£{cumulative_cash_flow:,.2f}"

                # Insert row in table
                row_id = self.tree.insert("", "end", values=(year, cash_flow_display, cumulative_display))

                # Apply green font to positive cumulative cash flow values
                if cumulative_cash_flow >= 0:
                    self.tree.tag_configure("positive", foreground="green")
                    self.tree.item(row_id, tags=("positive",))

                # Collect data for chart
                cash_flow_data.append((year, cumulative_cash_flow))

                year += 1

            # Store data for chart
            self.cash_flow_data = cash_flow_data

            # Clear the previous result
            self.result_label.config(text="")

            # Plot the chart
            self.plot_chart()

        except ValueError:
            messagebox.showerror("Input Error", "Please enter valid numbers for Initial Investment and Annual Net Benefits.")

    def calculate_payback_period(self):
        """Calculate the Payback Period based on cumulative cash flows."""
        try:
            cumulative_cash_flow = -self.initial_investment  # Start with negative initial investment
            year = 1
            max_years = 50  # Optional limit for payback period calculation

            while cumulative_cash_flow < 0 and year <= max_years:
                cumulative_cash_flow += self.annual_cash_flow
                if cumulative_cash_flow >= 0:
                    # Calculate the fractional year for the payback
                    previous_cumulative_cash_flow = cumulative_cash_flow - self.annual_cash_flow
                    remaining_cash_needed = -previous_cumulative_cash_flow
                    fraction_of_year = remaining_cash_needed / self.annual_cash_flow
                    payback_period_years = year - 1 + fraction_of_year
                    payback_years = int(payback_period_years)
                    payback_months = int(round((payback_period_years - payback_years) * 12))

                    # Display the result
                    self.result_label.config(
                        text=f"Payback Period: {payback_years} years and {payback_months} months"
                    )
                    return

                year += 1

            # If the cumulative cash flow never reaches the initial investment, show an error
            self.result_label.config(text="The project does not pay back within the specified period.")

        except (AttributeError, TypeError):
            messagebox.showerror("Error", "Please click 'Update' after entering values to initialize the table.")

    def plot_chart(self):
        """Generate and display the Cumulative Cash Flow chart."""
        try:
            # Extract data for plotting
            years = [data[0] for data in self.cash_flow_data]
            cumulative_cash_flows = [data[1] for data in self.cash_flow_data]

            # Clear previous plot
            self.ax.clear()

            # Plot Cumulative Cash Flow
            self.ax.plot(years, cumulative_cash_flows, marker='o', linestyle='-', color='blue', label='Cumulative Cash Flow (£)')

            # Identify Break-Even Point
            breakeven_year = None
            breakeven_cash_flow = None
            for i, cash_flow in enumerate(cumulative_cash_flows):
                if cash_flow >= 0:
                    breakeven_year = years[i]
                    breakeven_cash_flow = cash_flow
                    break

            if breakeven_year is not None:
                self.ax.plot(breakeven_year, breakeven_cash_flow, marker='o', color='green', label='Break-Even Point')
                self.ax.annotate(f'BE Point\nYear {breakeven_year}\n£{breakeven_cash_flow:,.2f}',
                                 xy=(breakeven_year, breakeven_cash_flow),
                                 xytext=(breakeven_year + 2, breakeven_cash_flow + self.initial_investment * 0.1),
                                 arrowprops=dict(facecolor='black', shrink=0.05),
                                 fontsize=10,
                                 horizontalalignment='left')

            # Add labels and title
            self.ax.set_title("Cumulative Cash Flow Over Time")
            self.ax.set_xlabel("Year")
            self.ax.set_ylabel("£")
            self.ax.legend()
            self.ax.grid(True)

            # Update the canvas with the new plot
            self.canvas.draw()

        except Exception as e:
            messagebox.showerror("Plot Error", f"An error occurred while plotting the chart:\n{e}")

    def download_to_excel(self):
        """Download the table data to an Excel file."""
        try:
            # Prepare data for the Excel file
            data = []
            for row in self.tree.get_children():
                values = self.tree.item(row)["values"]
                data.append(values)

            # Convert to DataFrame and save as Excel
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            file_path = os.path.join(desktop_path, "payback_period_calculation.xlsx")
            df = pd.DataFrame(data, columns=["Year", "Cash Flow", "Cumulative Cash Flow"])
            df.to_excel(file_path, index=False)

            # Confirmation message
            messagebox.showinfo("Download Complete", f"Table data has been saved to {file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel file: {e}")

    def download_chart(self):
        """Download the Payback Period chart as an image."""
        try:
            # Open a file dialog to choose save location
            file_path = filedialog.asksaveasfilename(
                defaultextension=".png",
                filetypes=[("PNG files", "*.png"), ("All files", "*.*")],
                title="Save Chart As"
            )
            if file_path:
                self.figure.savefig(file_path)
                messagebox.showinfo("Download Successful", f"Chart has been saved to {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save chart: {e}")

# Initialize the PaybackPeriodApp with the new frame
payback_app = PaybackPeriodApp(payback_period_frame)

# ----------------------------------------
# Fourth Tab: NPV Calculator
# ----------------------------------------

# Add a new tab for 'NPV Calculator'
npv_calculator_frame = ttk.Frame(notebook)
notebook.add(npv_calculator_frame, text='NPV Calculator')

# Define the NPV Calculator Class
class NPVCalculatorApp:
    def __init__(self, parent):
        self.parent = parent

        # Title Label
        ttk.Label(self.parent, text="Net Present Value (NPV) Calculator", font=("Helvetica", 16)).pack(pady=10)

        # Inputs Frame
        input_frame = ttk.Frame(self.parent)
        input_frame.pack(pady=10)

        # Discount Rate Input
        ttk.Label(input_frame, text="Discount Rate (%): ").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        self.discount_rate_entry = ttk.Entry(input_frame, width=15)
        self.discount_rate_entry.grid(row=0, column=1, padx=5, pady=5)
        self.discount_rate_entry.insert(0, "10")  # Default Data

        # Initial Investment Input
        ttk.Label(input_frame, text="Initial Investment (£): ").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        self.initial_investment_entry = ttk.Entry(input_frame, width=15)
        self.initial_investment_entry.grid(row=1, column=1, padx=5, pady=5)
        self.initial_investment_entry.insert(0, "10000")  # Default Data

        # Instruction Label next to Initial Investment
        ttk.Label(input_frame, text="To be a positive number.").grid(row=1, column=2, padx=5, pady=5, sticky='w')

        # Cash Flows Input
        ttk.Label(input_frame, text="Cash Flows (£) (comma-separated): ").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        self.cash_flows_entry = ttk.Entry(input_frame, width=30)
        self.cash_flows_entry.grid(row=2, column=1, padx=5, pady=5)
        self.cash_flows_entry.insert(0, "3000, 3500, 4000, 4500, 5000")  # Default Data

        # Calculate Button
        ttk.Button(self.parent, text="Calculate NPV", command=self.calculate_npv).pack(pady=10)

        # Results Labels Frame
        self.results_frame = ttk.Frame(self.parent)
        self.results_frame.pack(pady=10)

        self.total_pv_label = ttk.Label(self.results_frame, text="", font=("Helvetica", 12))
        self.total_pv_label.pack(pady=5)

        self.initial_investment_label = ttk.Label(self.results_frame, text="", font=("Helvetica", 12))
        self.initial_investment_label.pack(pady=5)

        self.npv_label = ttk.Label(self.results_frame, text="", font=("Helvetica", 12))
        self.npv_label.pack(pady=5)

        # Table Frame for Detailed NPV Breakdown
        self.npv_table_frame = ttk.Frame(self.parent)
        self.npv_table_frame.pack(pady=10, padx=20, fill='both', expand=True)

        # Set up Treeview (Table) with Scrollbar
        self.tree = ttk.Treeview(self.npv_table_frame, columns=("Year", "Cash Flow (£)", "Discount Factor", "Present Value (£)"), show="headings")
        self.tree.heading("Year", text="Year")
        self.tree.heading("Cash Flow (£)", text="Cash Flow (£)")
        self.tree.heading("Discount Factor", text="Discount Factor")
        self.tree.heading("Present Value (£)", text="Present Value (£)")
        self.tree.column("Year", anchor="center", width=50)
        self.tree.column("Cash Flow (£)", anchor="center", width=120)
        self.tree.column("Discount Factor", anchor="center", width=120)
        self.tree.column("Present Value (£)", anchor="center", width=150)

        # Add vertical scrollbar
        scrollbar = ttk.Scrollbar(self.npv_table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        self.tree.grid(row=0, column=0)
        scrollbar.grid(row=0, column=1, sticky="ns")

        # Download Buttons Frame
        download_buttons_frame = ttk.Frame(self.parent)
        download_buttons_frame.pack(pady=10)

        # Download NPV Calculation to Excel
        ttk.Button(download_buttons_frame, text="Download NPV Calculation", command=self.download_npv).grid(row=0, column=0, padx=10)

        # Download Chart Button
        ttk.Button(download_buttons_frame, text="Download Chart", command=self.download_chart).grid(row=0, column=1, padx=10)

        # Chart Frame
        self.chart_frame = ttk.Frame(self.parent)
        self.chart_frame.pack(pady=10, padx=20, fill='both', expand=True)

        # Initialize matplotlib Figure and Canvas
        self.figure = Figure(figsize=(6, 4), dpi=100)
        self.ax = self.figure.add_subplot(111)
        self.ax.set_title("NPV Analysis")
        self.ax.set_xlabel("Year")
        self.ax.set_ylabel("£")

        self.canvas = FigureCanvasTkAgg(self.figure, master=self.chart_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(fill='both', expand=True)

    def calculate_npv(self):
        """Calculate the Net Present Value based on user input and populate the table and chart."""
        try:
            discount_rate = float(self.discount_rate_entry.get()) / 100
            initial_investment = float(self.initial_investment_entry.get())
            cash_flows = [float(cf.strip()) for cf in self.cash_flows_entry.get().split(',')]

            npv = 0
            cash_flow_list = []
            total_pv_benefits = 0  # Initialize Total PV of Benefits

            for i, cf in enumerate(cash_flows, start=1):
                discount_factor = 1 / ((1 + discount_rate) ** i)
                present_value = cf * discount_factor
                npv += present_value
                total_pv_benefits += present_value
                cash_flow_list.append({
                    "Year": i,
                    "Cash Flow (£)": f"£{cf:,.2f}",
                    "Discount Factor": f"{discount_factor:.4f}",
                    "Present Value (£)": f"£{present_value:,.2f}"
                })

            # Subtract Initial Investment
            npv -= initial_investment

            # Clear existing rows in the table
            for item in self.tree.get_children():
                self.tree.delete(item)

            # Insert new rows into the table
            for row in cash_flow_list:
                self.tree.insert("", "end", values=(
                    row["Year"],
                    row["Cash Flow (£)"],
                    row["Discount Factor"],
                    row["Present Value (£)"]
                ))

            # Display the results in labels
            self.total_pv_label.config(text=f"Total PV of Benefits: £{total_pv_benefits:,.2f}")
            self.initial_investment_label.config(text=f"Initial Investment (£): £{initial_investment:,.2f}")
            self.npv_label.config(text=f"NPV (£): £{npv:,.2f}")

            # Plot the NPV Analysis Chart
            self.plot_chart(cash_flow_list)

        except ValueError:
            messagebox.showerror("Input Error", "Please enter valid numerical values.")

    def plot_chart(self, cash_flow_list):
        """Generate and display the NPV Analysis chart."""
        try:
            # Extract data for plotting
            years = [item["Year"] for item in cash_flow_list]
            cash_flows = [float(item["Cash Flow (£)"].replace('£', '').replace(',', '')) for item in cash_flow_list]
            present_values = [float(item["Present Value (£)"].replace('£', '').replace(',', '')) for item in cash_flow_list]

            # Clear previous plot
            self.ax.clear()

            # Plot Cash Flows and Present Values
            bar_width = 0.35
            index = range(len(years))

            self.ax.bar(index, cash_flows, bar_width, label='Cash Flow (£)', color='skyblue')
            self.ax.bar([i + bar_width for i in index], present_values, bar_width, label='Present Value (£)', color='salmon')

            # Add labels and title
            self.ax.set_title("NPV Analysis")
            self.ax.set_xlabel("Year")
            self.ax.set_ylabel("£")
            self.ax.set_xticks([i + bar_width / 2 for i in index])
            self.ax.set_xticklabels(years)
            self.ax.legend()
            self.ax.grid(axis='y')

            # Annotate bars with values
            for i in index:
                self.ax.text(i, cash_flows[i] + max(cash_flows)*0.01, f"£{cash_flows[i]:,.2f}", ha='center', va='bottom', fontsize=8)
                self.ax.text(i + bar_width, present_values[i] + max(present_values)*0.01, f"£{present_values[i]:,.2f}", ha='center', va='bottom', fontsize=8)

            # Update the canvas with the new plot
            self.canvas.draw()

        except Exception as e:
            messagebox.showerror("Plot Error", f"An error occurred while plotting the chart:\n{e}")

    def download_npv(self):
        """Download the NPV calculation details to an Excel file."""
        try:
            discount_rate = float(self.discount_rate_entry.get()) / 100
            initial_investment = float(self.initial_investment_entry.get())
            cash_flows = [float(cf.strip()) for cf in self.cash_flows_entry.get().split(',')]

            npv = 0
            cash_flow_list = []
            total_pv_benefits = 0  # Initialize Total PV of Benefits

            for i, cf in enumerate(cash_flows, start=1):
                discount_factor = 1 / ((1 + discount_rate) ** i)
                present_value = cf * discount_factor
                npv += present_value
                total_pv_benefits += present_value
                cash_flow_list.append({
                    "Year": i,
                    "Cash Flow (£)": cf,
                    "Discount Factor": discount_factor,
                    "Present Value (£)": present_value
                })

            # Subtract Initial Investment
            npv -= initial_investment

            # Create DataFrame
            df_cash_flows = pd.DataFrame(cash_flow_list)
            df_summary = pd.DataFrame({
                "Initial Investment (£)": [initial_investment],
                "Total PV of Benefits (£)": [total_pv_benefits],
                "NPV (£)": [npv]
            })

            # Save to Excel with multiple sheets
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            file_path = os.path.join(desktop_path, "NPV_Calculation.xlsx")
            with pd.ExcelWriter(file_path) as writer:
                df_cash_flows.to_excel(writer, sheet_name='Cash Flows', index=False)
                df_summary.to_excel(writer, sheet_name='Summary', index=False)

            # Save the chart as an image and embed it into the Excel file
            chart_path = os.path.join(desktop_path, "npv_chart.png")
            self.figure.savefig(chart_path)

            # Append the chart image to the Excel file (requires openpyxl and Pillow)
            try:
                from openpyxl import load_workbook
                from openpyxl.drawing.image import Image as OpenpyxlImage

                wb = load_workbook(file_path)
                ws = wb.create_sheet(title='NPV Chart')
                img = OpenpyxlImage(chart_path)
                ws.add_image(img, 'A1')
                wb.save(file_path)
                # Remove the standalone chart image
                os.remove(chart_path)
            except ImportError:
                messagebox.showwarning("Optional Feature Missing",
                                       "To embed the chart into Excel, please install 'openpyxl' and 'Pillow' libraries.\n\nThe chart has been saved separately as 'npv_chart.png' on your Desktop.")
            except Exception as e:
                messagebox.showwarning("Chart Embedding Failed",
                                       f"An error occurred while embedding the chart into Excel:\n{e}\n\nThe chart has been saved separately as 'npv_chart.png' on your Desktop.")

            # Confirmation message
            messagebox.showinfo("Download Complete", f"NPV calculation has been saved to {file_path}")

        except ValueError:
            messagebox.showerror("Input Error", "Please enter valid numerical values.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel file: {e}")

    def download_chart(self):
        """Download the NPV chart as an image."""
        try:
            # Open a file dialog to choose save location
            file_path = filedialog.asksaveasfilename(
                defaultextension=".png",
                filetypes=[("PNG files", "*.png"), ("All files", "*.*")],
                title="Save Chart As"
            )
            if file_path:
                self.figure.savefig(file_path)
                messagebox.showinfo("Download Successful", f"Chart has been saved to {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save chart: {e}")

# Initialize the NPVCalculatorApp with the new frame
npv_app = NPVCalculatorApp(npv_calculator_frame)

# ----------------------------------------
# Fifth Tab: Break-Even Analysis Tool
# ----------------------------------------

# Add a new tab for 'Break-Even Analysis'
break_even_frame = ttk.Frame(notebook)
notebook.add(break_even_frame, text='Break-Even Analysis')

# Define the BreakEvenAnalysisApp class
class BreakEvenAnalysisApp:
    def __init__(self, parent):
        self.parent = parent

        # Title Label
        ttk.Label(self.parent, text="Break-Even Analysis Tool", font=("Helvetica", 16)).pack(pady=10)

        # Inputs Frame
        input_frame = ttk.Frame(self.parent)
        input_frame.pack(pady=10)

        # Fixed Costs Input
        ttk.Label(input_frame, text="Fixed Costs (£): ").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        self.fixed_costs_entry = ttk.Entry(input_frame, width=15)
        self.fixed_costs_entry.grid(row=0, column=1, padx=5, pady=5)
        self.fixed_costs_entry.insert(0, "5000")  # Default Data

        # Instruction Label next to Fixed Costs
        ttk.Label(input_frame, text="(Enter as a positive number)").grid(row=0, column=2, padx=5, pady=5, sticky='w')

        # Variable Cost per Unit Input
        ttk.Label(input_frame, text="Variable Cost per Unit (£): ").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        self.variable_cost_entry = ttk.Entry(input_frame, width=15)
        self.variable_cost_entry.grid(row=1, column=1, padx=5, pady=5)
        self.variable_cost_entry.insert(0, "20")  # Default Data

        # Instruction Label next to Variable Cost per Unit
        ttk.Label(input_frame, text="(Enter as a positive number)").grid(row=1, column=2, padx=5, pady=5, sticky='w')

        # Sales Price per Unit Input
        ttk.Label(input_frame, text="Sales Price per Unit (£): ").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        self.sales_price_entry = ttk.Entry(input_frame, width=15)
        self.sales_price_entry.grid(row=2, column=1, padx=5, pady=5)
        self.sales_price_entry.insert(0, "50")  # Default Data

        # Instruction Label next to Sales Price per Unit
        ttk.Label(input_frame, text="(Enter as a positive number)").grid(row=2, column=2, padx=5, pady=5, sticky='w')

        # Calculate Button
        ttk.Button(self.parent, text="Calculate Break-Even", command=self.calculate_break_even).pack(pady=10)

        # Results Frame
        self.results_frame = ttk.Frame(self.parent)
        self.results_frame.pack(pady=10)
        self.breakeven_units_label = ttk.Label(self.results_frame, text="", font=("Helvetica", 12))
        self.breakeven_units_label.pack(pady=5)
        self.breakeven_revenue_label = ttk.Label(self.results_frame, text="", font=("Helvetica", 12))
        self.breakeven_revenue_label.pack(pady=5)

        # Chart Frame
        self.chart_frame = ttk.Frame(self.parent)
        self.chart_frame.pack(pady=10, padx=20, fill='both', expand=True)

        # Initialize matplotlib Figure and Canvas
        self.figure = Figure(figsize=(6, 4), dpi=100)
        self.ax = self.figure.add_subplot(111)
        self.ax.set_title("Break-Even Analysis")
        self.ax.set_xlabel("Units Sold")
        self.ax.set_ylabel("£")

        self.canvas = FigureCanvasTkAgg(self.figure, master=self.chart_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(fill='both', expand=True)

        # Download Buttons Frame for Chart
        download_buttons_frame = ttk.Frame(self.parent)
        download_buttons_frame.pack(pady=10)

        # Download Break-Even Analysis to Excel
        ttk.Button(download_buttons_frame, text="Download Break-Even Analysis", command=self.download_break_even).grid(row=0, column=0, padx=10)

        # Download Chart Button
        ttk.Button(download_buttons_frame, text="Download Chart", command=self.download_chart).grid(row=0, column=1, padx=10)

    def calculate_break_even(self):
        """Calculate the Break-Even Point based on user input and generate a chart."""
        try:
            fixed_costs = float(self.fixed_costs_entry.get())
            variable_cost = float(self.variable_cost_entry.get())
            sales_price = float(self.sales_price_entry.get())

            if sales_price <= variable_cost:
                messagebox.showerror("Input Error", "Sales Price per Unit must be greater than Variable Cost per Unit.")
                return

            # Calculate Break-Even Point in Units
            breakeven_units = fixed_costs / (sales_price - variable_cost)
            breakeven_units = round(breakeven_units, 2)

            # Calculate Break-Even Point in Revenue
            breakeven_revenue = breakeven_units * sales_price
            breakeven_revenue = round(breakeven_revenue, 2)

            # Display the results
            self.breakeven_units_label.config(text=f"Break-Even Point: {breakeven_units} units")
            self.breakeven_revenue_label.config(text=f"Break-Even Revenue: £{breakeven_revenue:,.2f}")

            # Generate data for chart
            max_units = int(breakeven_units * 1.5)  # Extend to 150% of break-even units for better visualization
            units = list(range(0, max_units + 1))
            total_costs = [fixed_costs + (variable_cost * u) for u in units]
            total_revenues = [sales_price * u for u in units]

            # Clear previous plot
            self.ax.clear()

            # Plot Total Costs and Total Revenues
            self.ax.plot(units, total_costs, label='Total Costs (£)', color='red', linewidth=2)
            self.ax.plot(units, total_revenues, label='Total Revenues (£)', color='green', linewidth=2)

            # Plot Break-Even Point
            self.ax.plot(breakeven_units, breakeven_revenue, 'bo', label='Break-Even Point')
            self.ax.annotate(f'BE Point\n({breakeven_units}, £{breakeven_revenue:,.2f})',
                             xy=(breakeven_units, breakeven_revenue),
                             xytext=(breakeven_units + max_units * 0.05, breakeven_revenue),
                             arrowprops=dict(facecolor='black', shrink=0.05),
                             fontsize=10,
                             horizontalalignment='left')

            # Add labels and title
            self.ax.set_title("Break-Even Analysis")
            self.ax.set_xlabel("Units Sold")
            self.ax.set_ylabel("£")
            self.ax.legend()
            self.ax.grid(True)

            # Update the canvas with the new plot
            self.canvas.draw()

        except ValueError:
            messagebox.showerror("Input Error", "Please enter valid numerical values.")

    def download_break_even(self):
        """Download the Break-Even Analysis to an Excel file."""
        try:
            fixed_costs = float(self.fixed_costs_entry.get())
            variable_cost = float(self.variable_cost_entry.get())
            sales_price = float(self.sales_price_entry.get())

            if sales_price <= variable_cost:
                messagebox.showerror("Input Error", "Sales Price per Unit must be greater than Variable Cost per Unit.")
                return

            # Calculate Break-Even Point in Units
            breakeven_units = fixed_costs / (sales_price - variable_cost)
            breakeven_units = round(breakeven_units, 2)

            # Calculate Break-Even Point in Revenue
            breakeven_revenue = breakeven_units * sales_price
            breakeven_revenue = round(breakeven_revenue, 2)

            # Prepare data for Excel
            data = {
                "Fixed Costs (£)": [fixed_costs],
                "Variable Cost per Unit (£)": [variable_cost],
                "Sales Price per Unit (£)": [sales_price],
                "Break-Even Point (Units)": [breakeven_units],
                "Break-Even Revenue (£)": [breakeven_revenue]
            }

            df = pd.DataFrame(data)
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            file_path = os.path.join(desktop_path, "break_even_analysis.xlsx")
            df.to_excel(file_path, index=False)

            # Save the chart as an image and embed it into the Excel file
            chart_path = os.path.join(desktop_path, "break_even_chart.png")
            self.figure.savefig(chart_path)

            # Append the chart image to the Excel file (requires openpyxl and Pillow)
            try:
                from openpyxl import load_workbook
                from openpyxl.drawing.image import Image as OpenpyxlImage

                wb = load_workbook(file_path)
                ws = wb.create_sheet(title='Break-Even Chart')
                img = OpenpyxlImage(chart_path)
                ws.add_image(img, 'A1')
                wb.save(file_path)
                # Remove the standalone chart image
                os.remove(chart_path)
            except ImportError:
                messagebox.showwarning("Optional Feature Missing",
                                       "To embed the chart into Excel, please install 'openpyxl' and 'Pillow' libraries.\n\nThe chart has been saved separately as 'break_even_chart.png' on your Desktop.")
            except Exception as e:
                messagebox.showwarning("Chart Embedding Failed",
                                       f"An error occurred while embedding the chart into Excel:\n{e}\n\nThe chart has been saved separately as 'break_even_chart.png' on your Desktop.")

            # Confirmation message
            messagebox.showinfo("Download Complete", f"Break-Even Analysis has been saved to {file_path}")

        except ValueError:
            messagebox.showerror("Input Error", "Please enter valid numerical values.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel file: {e}")

    def download_chart(self):
        """Download the Break-Even chart as an image."""
        try:
            # Open a file dialog to choose save location
            file_path = filedialog.asksaveasfilename(
                defaultextension=".png",
                filetypes=[("PNG files", "*.png"), ("All files", "*.*")],
                title="Save Chart As"
            )
            if file_path:
                self.figure.savefig(file_path)
                messagebox.showinfo("Download Successful", f"Chart has been saved to {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save chart: {e}")

# Initialize the BreakEvenAnalysisApp with the new frame
break_even_app = BreakEvenAnalysisApp(break_even_frame)

# ----------------------------------------
# Start the Tkinter event loop
# ----------------------------------------

root.mainloop()
